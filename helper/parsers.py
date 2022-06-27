# implementirati klasu za parsiranje fajlova sa podacima datih u primeru

import csv
import openpyxl
from openpyxl import load_workbook

class FileParser(object):
    # klasa treba da sadrzi konstruktor u kome se definisu atributi koji
    # cuvaju putanju do ulaznog fajla kao i tip fajla. tip moze biti
    # prosledjen prilikom instanciranja ili automatski prepoznat pomocu ekstenzije

    # klasa treba da sadrzi parser metodu za svaki tip ulazne datoteke
    # metoda 1: excel_to_list
    # metoda 2: csv_to_list
    # metoda treba da otvori datoteku i izdvoji sve vrednosti signala i
    # vrati listu tih vrednosti
    # koristiti python bilioteke za rad sa csv i excel fajlovima (moduli csv, openpyxl...)
    # prilikom ucitavanja datoteka za svaku metodu implementirati error handling,
    # da ukoliko fajl ne postoji dobijemo o tome poruku i u tom slucaju metoda vraca praznu listu

    # klasa treba da sadrzi metodu data_to_list koja ce u zavisnosti od atributa
    # koji ukazuje na tip fajla pozvati odgovarajucu metodu klase i vratiti listu vrednosti


    def __init__(self, path, ext):
        self.path = path
        self.ext = ext


    def excel_to_list(self):
        print("excel_to_list")
        #if os.path.exists(file)
        try:
            wb = load_workbook(filename = self.path)
            sheet = wb["Sheet1"]
            cells = []
            for row in sheet.iter_rows(min_row=2, min_col=2, max_row=2):
                for cell in row:
                    cells.append(cell.value)
            return cells
        except FileNotFoundError:      
            print("No such file or directory")
            return []


    def csv_to_list(self):
        print("csv_to_list")
        try:        
            with open(self.path, "r") as file:
                csvreader = csv.reader(file) # read the CSV file
                _ = next(csvreader) # skip header
                # list comprehension
                rows = [float(row[1]) for row in csvreader] # iterate through the csvreader object 
                return(rows)
        except FileNotFoundError:  
            print("No such file or directory")
            return []


    def data_to_list(self):
        if self.ext == "csv":
            return self.csv_to_list()
        else:
            return self.excel_to_list()


# implementirati drugo klasu FileParserExtended koja nasledjuje FileParser klasu i definisati
# metode excel_to_dict i csv_to_dict koje vracaju recnike gde su parovi {naziv_sinala: vrednost}


class FileParserExtended(FileParser):


    def excel_to_dict(self):
        print("excel_to_dict")
        try:
            wb = load_workbook(filename = self.path)
            sheet = wb["Sheet1"]
            singnals_dict = {}
            for col in sheet.iter_cols(min_row=1, min_col=2, max_row=2):
                    #print(col[0].value)
                    #print(col[1].value)
                    # putting pairs in dictionary during an iteration
                    singnals_dict[col[0].value] = col[1].value 
            return singnals_dict
        except FileNotFoundError:   
            print("No such file or directory")
            return {}


    def csv_to_dict(self):
        print("csv_to_dict")
        try:        
            with open(self.path, "r") as file:
                csvreader = csv.reader(file) # read the CSV file
                _ = next(csvreader) # skip header
                # dictionary comprehension
                return {row[0]: float(row[1]) for row in csvreader}
        except FileNotFoundError:  
            print("No such file or directory")
            return {}

    
    def data_to_dict(self):
        if self.ext == "csv":
            return self.csv_to_dict()
        else:
            return self.excel_to_dict()

